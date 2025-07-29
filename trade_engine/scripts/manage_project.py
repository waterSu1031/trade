import os
import ast
import pandas as pd
import chardet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 🔍 호출된 함수 이름 추출 도우미
def get_call_name(n):
    if isinstance(n, ast.Name):
        return n.id
    elif isinstance(n, ast.Attribute):
        base = get_call_name(n.value)
        return f"{base}.{n.attr}" if base else n.attr
    return ""

# 1️⃣ Docstring 추출 함수
def extract_docstrings(node):
    doc_info = []
    class Visitor(ast.NodeVisitor):
        def __init__(self):
            self.current_class = None

        def visit_ClassDef(self, class_node):
            doc_info.append(("Class", class_node.name, self.current_class or "", ast.get_docstring(class_node) or "", class_node.lineno))
            prev = self.current_class
            self.current_class = class_node.name
            for child in class_node.body:
                self.visit(child)
            self.current_class = prev

        def visit_FunctionDef(self, func_node):
            doc_info.append(("Function", func_node.name, self.current_class or "", ast.get_docstring(func_node) or "", func_node.lineno))

    Visitor().visit(node)
    return doc_info

# 2️⃣ Calls 추출 함수
def extract_calls(node, project_symbols, local_symbols):
    call_info = {}
    class Visitor(ast.NodeVisitor):
        def __init__(self):
            self.current_func = None

        def visit_FunctionDef(self, func_node):
            self.current_func = func_node.name
            raw_calls = [get_call_name(n.func) for n in ast.walk(func_node) if isinstance(n, ast.Call)]
            filtered = [c for c in raw_calls if c.split('.')[0] in project_symbols and c.split('.')[0] not in local_symbols]
            call_info[self.current_func] = ", ".join(sorted(set(filtered)))

    Visitor().visit(node)
    return call_info

# 3️⃣ Comment 추출 함수 (# @desc: 설명)
def extract_comment_map(lines, lineno_map):
    comments = {}
    for name, lineno in lineno_map.items():
        for i in range(lineno - 2, max(lineno - 10, -1), -1):
            line = lines[i].strip()
            if line.startswith("# @desc:"):
                comments[name] = line.replace("# @desc:", "").strip()
                break
            elif line.startswith("#") or not line:
                continue
            else:
                break
    return comments

# 4️⃣ 파일 경로를 레벨로 변환

def extract_levels(base_dir, full_path):
    rel_path = os.path.relpath(full_path, base_dir).replace("\\", "/")
    parts = rel_path.split("/")
    while len(parts) < 5:
        parts.append("")
    parts.append("")
    return parts

# 🔁 통합 구조 분석 함수

def extract_project_structure(base_dir):
    ignore_dirs = {'.venv', '__pycache__', '.git'}
    records = []
    project_symbols = extract_project_symbols(base_dir)

    for root, dirs, files in os.walk(base_dir):
        dirs[:] = [d for d in dirs if d not in ignore_dirs]

        for file in files:
            if not file.endswith('.py'):
                continue

            full_path = os.path.join(root, file)
            local_symbols = extract_local_symbols(full_path)

            with open(full_path, 'rb') as f:
                raw = f.read()
            encoding = chardet.detect(raw)['encoding'] or 'utf-8'
            content = raw.decode(encoding, errors='ignore')
            lines = content.splitlines()

            try:
                node = ast.parse(content, filename=full_path)
            except SyntaxError:
                continue

            doc_items = extract_docstrings(node)
            call_map = extract_calls(node, project_symbols, local_symbols)
            lineno_map = {name: lineno for _, name, _, _, lineno in doc_items}
            comment_map = extract_comment_map(lines, lineno_map)

            levels = extract_levels(base_dir, full_path)

            for typ, name, parent, doc, _ in doc_items:
                records.append({
                    "Level_1": levels[0],
                    "Level_2": levels[1],
                    "Level_3": levels[2],
                    "Level_4": levels[3],
                    "Level_5": levels[4],
                    "Type": typ,
                    "File": file,
                    "Name": name,
                    "Docstring": doc,
                    "Comment": comment_map.get(name, ""),
                    "Calls": call_map.get(name, "")
                })

    return pd.DataFrame(records)

# 🔧 심볼 추출 함수

def extract_project_symbols(base_dir):
    symbols = set()
    for root, _, files in os.walk(base_dir):
        if any(ignored in root for ignored in ['.venv', '__pycache__', '.git']):
            continue
        for file in files:
            if file.endswith('.py'):
                path = os.path.join(root, file)
                try:
                    with open(path, 'rb') as f:
                        raw = f.read()
                    encoding = chardet.detect(raw)['encoding'] or 'utf-8'
                    content = raw.decode(encoding, errors='ignore')
                    node = ast.parse(content)
                    for n in ast.walk(node):
                        if isinstance(n, (ast.FunctionDef, ast.ClassDef)):
                            symbols.add(n.name)
                except Exception:
                    continue
    return symbols

def extract_local_symbols(file_path):
    local = set()
    try:
        with open(file_path, 'rb') as f:
            raw = f.read()
        encoding = chardet.detect(raw)['encoding'] or 'utf-8'
        content = raw.decode(encoding, errors='ignore')
        node = ast.parse(content)
        for n in ast.walk(node):
            if isinstance(n, (ast.FunctionDef, ast.ClassDef)):
                local.add(n.name)
    except Exception:
        pass
    return local

# 💾 엑셀 저장

def save_structure_to_excel(base_dir, output_file='project_structure.xlsx'):
    df = extract_project_structure(base_dir)
    col_order = [
        "Level_1", "Level_2", "Level_3", "Level_4", "Level_5",
        "File", "Type", "Name", "Docstring", "Comment", "Calls"
    ]
    df = df[df.columns.intersection(col_order)]
    df.to_excel(output_file, index=False)
    format_excel_design(output_file)
    print(f"✅[SUCCESS] Excel saved to: {output_file}")

# 📐 포맷팅 함수

def format_excel_design(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # 헤더 위치 매핑
    headers = {cell.value: cell.column for cell in ws[1]}
    level_cols = [headers.get(f"Level_{i}") for i in range(1, 6)]
    file_col = headers.get("File")
    type_col = headers.get("Type")
    name_col = headers.get("Name")
    doc_col = headers.get("Docstring")
    comment_col = headers.get("Comment")

    # 1️⃣ 병합 처리: Level_1~5, File
    merge_cols = level_cols + [file_col]
    for col_idx in merge_cols:
        if not col_idx:
            continue
        start = 2
        prev_val = ws.cell(row=2, column=col_idx).value
        for row in range(3, ws.max_row + 2):
            curr_val = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else None
            if curr_val != prev_val:
                if row - start > 1:
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                    ws.cell(row=start, column=col_idx).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                start = row
                prev_val = curr_val

    # 2️⃣ 기본 셀 정렬: 세로 중앙
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # 3️⃣ 좌우 정렬: Level, Type = 중앙 / File = 왼쪽
    for r in range(2, ws.max_row + 1):
        for c in level_cols:
            if c:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if type_col:
            ws.cell(row=r, column=type_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if file_col:
            ws.cell(row=r, column=file_col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 4️⃣ 열 너비 자동 + Docstring/Comment 너비 조정
    name_width = 20  # 기본 fallback
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=10)
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        if col_idx == name_col:
            name_width = adjusted_width

    if doc_col:
        ws.column_dimensions[get_column_letter(doc_col)].width = name_width * 1.5
    if comment_col:
        ws.column_dimensions[get_column_letter(comment_col)].width = name_width * 1.5

    wb.save(excel_path)
    wb.close()

# ▶️ 실행 예시
if __name__ == '__main__':
    base_dir = os.path.abspath('..')
    save_structure_to_excel(base_dir)
